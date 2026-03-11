from __future__ import annotations

import base64
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
import re

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

DEFAULT_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1qWFq1BTFnsxuKLnhprXz4N0FYfmL1lVf-hG_ldSf1jQ/edit?usp=sharing"
)

MATCH_COLUMNS = [
    "Zona",
    "Día",
    "Hora",
    "Complejo",
    "Código 1",
    "Pareja 1",
    "Código 2",
    "Pareja 2",
    "Set1 P1",
    "Set1 P2",
    "Set2 P1",
    "Set2 P2",
    "Set3 P1",
    "Set3 P2",
    "Estado",
]

BACKGROUND_IMAGE_PATH = Path("assets/logo.png")


def extract_sheet_id(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        raise ValueError("No pude detectar el ID del Google Sheet en la URL.")
    return match.group(1)


def to_export_url(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


@st.cache_data(ttl=60)
def download_workbook_bytes(sheet_url: str) -> bytes:
    sheet_id = extract_sheet_id(sheet_url)
    response = requests.get(to_export_url(sheet_id), timeout=30)
    response.raise_for_status()
    return response.content


@st.cache_data(ttl=60)
def get_visible_category_sheets(workbook_bytes: bytes) -> list[str]:
    workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True, data_only=True)
    return [
        worksheet.title
        for worksheet in workbook.worksheets
        if worksheet.sheet_state == "visible" and "Zonas" in worksheet.title
    ]


def clean(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).replace("\n", " ").strip()
    return " ".join(text.split())


def format_score_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def format_hour(value: object) -> str:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, date):
        return ""
    text = clean(value)
    if len(text) >= 8 and text.count(":") == 2:
        return text[:5]
    return text


def load_background_data_uri(image_path: Path) -> str:
    if not image_path.exists():
        return ""
    encoded = base64.b64encode(image_path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def parse_matches_from_worksheet(worksheet) -> pd.DataFrame:
    def cell_value(values: tuple[object, ...], index: int) -> object:
        if index < len(values):
            return values[index]
        return ""

    rows: list[dict[str, str]] = []
    current_zone = ""

    for row_index, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if worksheet.row_dimensions[row_index].hidden:
            continue

        first_col = clean(cell_value(row_values, 0))

        if first_col.upper().startswith("ZONA"):
            current_zone = first_col
            continue

        if not current_zone:
            continue

        code_1 = first_col
        team_1 = clean(cell_value(row_values, 1))
        code_2 = clean(cell_value(row_values, 3))
        team_2 = clean(cell_value(row_values, 4))
        day_label = clean(cell_value(row_values, 6))

        if day_label.lower() == "día":
            continue

        looks_like_match = bool(code_1 and code_2 and team_1 and team_2 and day_label)
        if not looks_like_match:
            continue

        set_1_a = format_score_value(cell_value(row_values, 9))
        set_1_b = format_score_value(cell_value(row_values, 10))
        set_2_a = format_score_value(cell_value(row_values, 11))
        set_2_b = format_score_value(cell_value(row_values, 12))
        set_3_a = format_score_value(cell_value(row_values, 13))
        set_3_b = format_score_value(cell_value(row_values, 14))
        status = clean(cell_value(row_values, 22))

        rows.append(
            {
                "Zona": current_zone,
                "Día": day_label,
                "Hora": format_hour(cell_value(row_values, 7)),
                "Complejo": clean(cell_value(row_values, 8)),
                "Código 1": code_1,
                "Pareja 1": team_1,
                "Código 2": code_2,
                "Pareja 2": team_2,
                "Set1 P1": set_1_a,
                "Set1 P2": set_1_b,
                "Set2 P1": set_2_a,
                "Set2 P2": set_2_b,
                "Set3 P1": set_3_a,
                "Set3 P2": set_3_b,
                "Estado": status or "No Jugado",
            }
        )

    return pd.DataFrame(rows, columns=MATCH_COLUMNS)


def parse_zone_matches(workbook_bytes: bytes, category_sheet: str) -> pd.DataFrame:
    workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    worksheet = workbook[category_sheet]
    return parse_matches_from_worksheet(worksheet)


@st.cache_data(ttl=60)
def parse_all_visible_matches(workbook_bytes: bytes, category_sheets: tuple[str, ...]) -> pd.DataFrame:
    workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    all_frames: list[pd.DataFrame] = []

    for category_sheet in category_sheets:
        if category_sheet not in workbook.sheetnames:
            continue

        parsed = parse_matches_from_worksheet(workbook[category_sheet])
        if parsed.empty:
            continue

        parsed["Categoría"] = display_category_name(category_sheet)
        all_frames.append(parsed)

    if not all_frames:
        return pd.DataFrame(columns=[*MATCH_COLUMNS, "Categoría"])

    return pd.concat(all_frames, ignore_index=True)


def display_category_name(category_sheet: str) -> str:
    return re.sub(r"\s*Zonas\s*", "", category_sheet, flags=re.IGNORECASE).strip()


def build_score_text(match: pd.Series) -> str:
    sets = [
        (match.get("Set1 P1", ""), match.get("Set1 P2", "")),
        (match.get("Set2 P1", ""), match.get("Set2 P2", "")),
        (match.get("Set3 P1", ""), match.get("Set3 P2", "")),
    ]
    parts = [f"{a}-{b}" for a, b in sets if a != "" and b != ""]
    if parts:
        return " | ".join(parts)
    return str(match.get("Estado", "No Jugado"))


def is_match_played(match: pd.Series) -> bool:
    status = clean(match.get("Estado", "")).lower()
    if status and status not in {"no jugado", "0", "-"}:
        return True

    numeric_scores: list[float] = []
    for column in ["Set1 P1", "Set1 P2", "Set2 P1", "Set2 P2", "Set3 P1", "Set3 P2"]:
        score_text = clean(match.get(column, ""))
        if score_text == "":
            continue
        try:
            numeric_scores.append(float(score_text))
        except ValueError:
            return True

    if numeric_scores and any(score > 0 for score in numeric_scores):
        return True

    return False


def render_category_status_summary(matches: pd.DataFrame) -> None:
    if matches.empty or "Categoría" not in matches.columns:
        return

    summary = (
        matches.groupby("Categoría", as_index=False)["Jugado"]
        .agg(Total="count", Jugados="sum")
        .sort_values("Categoría")
    )
    summary["Faltan"] = summary["Total"] - summary["Jugados"]

    cards = []
    for _, row in summary.iterrows():
        cards.append(
            (
                f"<div class=\"summary-card\">"
                f"<div class=\"summary-title\">🏷️ {row['Categoría']}</div>"
                f"<div class=\"summary-stats\">✅ Jugados: <b>{int(row['Jugados'])}</b> · "
                f"⏳ Faltan: <b>{int(row['Faltan'])}</b></div>"
                "</div>"
            )
        )

    st.markdown(f"<div class='summary-grid'>{''.join(cards)}</div>", unsafe_allow_html=True)


def render_match_cards(matches: pd.DataFrame) -> None:
    for _, match in matches.iterrows():
        st.markdown(
            f"""
            <div class="match-card">
                <div class="match-header">📅 {match['Día']} · 🕒 {match['Hora']} · 📍 {match['Complejo']}</div>
                <div class="teams-row">
                    <span class="team-tag">{match['Código 1']}</span>
                    <span class="team-name">{match['Pareja 1']}</span>
                </div>
                <div class="teams-row">
                    <span class="team-tag">{match['Código 2']}</span>
                    <span class="team-name">{match['Pareja 2']}</span>
                </div>
                <div class="match-score">🎾 Resultado: {build_score_text(match)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def main() -> None:
    st.set_page_config(page_title="Resultados JPQ", page_icon="🎾", layout="centered")

    background_data_uri = load_background_data_uri(BACKGROUND_IMAGE_PATH)

    components.html(
        """
        <script>
            setTimeout(function () {
                window.parent.location.reload();
            }, 240000);
        </script>
        """,
        height=0,
    )

    background_css = """
        [data-testid="stAppViewContainer"] {
            position: relative;
            background:
                radial-gradient(circle at 8% 12%, rgba(59, 130, 246, 0.12) 0, rgba(59, 130, 246, 0.12) 90px, transparent 91px),
                radial-gradient(circle at 92% 18%, rgba(16, 185, 129, 0.10) 0, rgba(16, 185, 129, 0.10) 110px, transparent 111px),
                radial-gradient(circle at 85% 82%, rgba(56, 189, 248, 0.12) 0, rgba(56, 189, 248, 0.12) 75px, transparent 76px),
                radial-gradient(circle at 12% 86%, rgba(20, 184, 166, 0.10) 0, rgba(20, 184, 166, 0.10) 95px, transparent 96px),
                linear-gradient(180deg, rgba(250, 251, 255, 0.97), rgba(245, 252, 249, 0.97));
        }
    """

    logo_css = ""
    if background_data_uri:
        background_css = f"""
            [data-testid="stAppViewContainer"] {{
                position: relative;
                background-image:
                    radial-gradient(circle at 8% 12%, rgba(59, 130, 246, 0.12) 0, rgba(59, 130, 246, 0.12) 90px, transparent 91px),
                    radial-gradient(circle at 92% 18%, rgba(16, 185, 129, 0.10) 0, rgba(16, 185, 129, 0.10) 110px, transparent 111px),
                    radial-gradient(circle at 85% 82%, rgba(56, 189, 248, 0.12) 0, rgba(56, 189, 248, 0.12) 75px, transparent 76px),
                    radial-gradient(circle at 12% 86%, rgba(20, 184, 166, 0.10) 0, rgba(20, 184, 166, 0.10) 95px, transparent 96px),
                    linear-gradient(rgba(255, 255, 255, 0.95), rgba(255, 255, 255, 0.95)),
                    url('{background_data_uri}');
                background-repeat: no-repeat;
                background-position: center;
                background-size: min(80vw, 500px);
                background-attachment: fixed;
            }}
        """
        logo_css = f"""
            <div class="logo-badge">
                <img src="{background_data_uri}" alt="JPQ logo" />
            </div>
        """

    css = """
        <style>
            .block-container {
                padding-top: 1.2rem;
                padding-bottom: 2rem;
                max-width: 980px;
                margin-left: auto;
                margin-right: auto;
            }
            .stApp h1 {
                letter-spacing: 0.3px;
                font-weight: 800;
                color: #0f172a;
                text-align: center;
                width: 100%;
            }
            .stSubheader {
                text-align: center;
            }
            .stApp, .stApp p, .stApp label, .stApp div {
                color: #0f172a;
            }
            .match-card {
                border: 1px solid rgba(148, 163, 184, 0.28);
                border-radius: 16px;
                padding: 1rem;
                margin-bottom: 0.8rem;
                background: rgba(255, 255, 255, 0.68);
                backdrop-filter: blur(8px);
                box-shadow: 0 10px 24px rgba(15, 23, 42, 0.08);
            }
            .match-header {
                font-weight: 700;
                margin-bottom: 0.5rem;
                color: #0f172a;
                text-align: center;
            }
            .teams-row {
                display: flex;
                gap: 0.55rem;
                align-items: center;
                justify-content: center;
                margin-bottom: 0.2rem;
            }
            .team-tag {
                min-width: 2.3rem;
                text-align: center;
                padding: 0.15rem 0.4rem;
                border-radius: 999px;
                border: 1px solid rgba(59, 130, 246, 0.35);
                font-size: 0.8rem;
                background: rgba(59, 130, 246, 0.12);
                color: #1d4ed8;
                font-weight: 700;
            }
            .team-name {font-weight: 600; color: #0f172a;}
            .match-score {
                margin-top: 0.5rem;
                font-size: 0.95rem;
                color: #0f766e;
                font-weight: 600;
                text-align: center;
            }
            .zone-title {
                margin-top: 1rem;
                margin-bottom: 0.55rem;
                font-weight: 800;
                color: #0f172a;
                text-align: center;
            }
            .logo-badge {
                position: fixed;
                right: 1rem;
                top: 0.8rem;
                z-index: 20;
                width: 58px;
                height: 58px;
                border-radius: 50%;
                padding: 6px;
                background: rgba(255,255,255,0.72);
                border: 1px solid rgba(148, 163, 184, 0.38);
                box-shadow: 0 8px 16px rgba(15, 23, 42, 0.14);
                backdrop-filter: blur(5px);
            }
            .logo-badge img {
                width: 100%;
                height: 100%;
                object-fit: contain;
            }
            .summary-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                gap: 0.55rem;
                margin-bottom: 0.8rem;
            }
            .summary-card {
                border-radius: 12px;
                border: 1px solid rgba(148, 163, 184, 0.35);
                background: rgba(255, 255, 255, 0.72);
                backdrop-filter: blur(6px);
                padding: 0.65rem 0.75rem;
            }
            .summary-title {
                font-weight: 800;
                color: #0f172a;
                margin-bottom: 0.2rem;
            }
            .summary-stats {
                color: #334155;
                font-size: 0.9rem;
                display: flex;
                align-items: center;
                justify-content: center;
                gap: 0.35rem;
                white-space: nowrap;
            }
            @media (max-width: 640px) {
                .block-container {padding-left: 0.65rem; padding-right: 0.65rem;}
                h1 {font-size: 1.45rem !important;}
                .match-card {
                    padding: 0.8rem;
                    border-radius: 14px;
                }
                .match-header {
                    font-size: 0.92rem;
                }
                .team-name {
                    font-size: 0.92rem;
                }
                .logo-badge {
                    width: 46px;
                    height: 46px;
                    right: 0.65rem;
                }
                .summary-stats {
                    font-size: 0.84rem;
                    gap: 0.25rem;
                }
            }
            __BACKGROUND_CSS__
        </style>
    """

    st.markdown(css.replace("__BACKGROUND_CSS__", background_css), unsafe_allow_html=True)
    if logo_css:
        st.markdown(logo_css, unsafe_allow_html=True)

    st.title("JPQ 1er Abierto 2026")

    with st.sidebar:
        st.header("Configuración")
        sheet_url = st.text_input("URL Google Sheet", value=DEFAULT_SHEET_URL)
        st.caption("Tip: necesitás permisos públicos de lectura para que funcione.")

    try:
        with st.spinner("Cargando categorías..."):
            workbook_bytes = download_workbook_bytes(sheet_url)
            categories = get_visible_category_sheets(workbook_bytes)
    except Exception as error:
        st.error(f"No pude leer la planilla: {error}")
        st.stop()

    if not categories:
        st.warning("No encontré pestañas visibles de tipo 'Zonas'.")
        st.stop()

    category_labels = {display_category_name(name): name for name in categories}
    category_option_labels = ["Todas las categorías"] + list(category_labels.keys())
    selected_label = st.selectbox("🏷️ Categoría", options=category_option_labels)

    all_matches = parse_all_visible_matches(workbook_bytes, tuple(categories))

    if all_matches.empty:
        st.info("No se detectaron partidos en la selección actual.")
        st.stop()

    matches = all_matches.copy()
    if selected_label != "Todas las categorías":
        matches = matches[matches["Categoría"] == selected_label].copy()

    if matches.empty:
        st.info("No se detectaron partidos en la selección actual.")
        st.stop()

    matches["Jugado"] = matches.apply(is_match_played, axis=1)
    render_category_status_summary(matches)

    all_zones = sorted(matches["Zona"].dropna().unique().tolist())
    selected_zone = st.selectbox("🗺️ Zona (opcional)", options=["Todas"] + all_zones)

    all_complexes = sorted(
        [complex_name for complex_name in matches["Complejo"].dropna().unique().tolist() if complex_name]
    )
    selected_complex = st.selectbox("📍 Complejo (opcional)", options=["Todos"] + all_complexes)

    surname_query = st.text_input("🔎 Buscar apellido", placeholder="Ej: PAREDES, MARTINEZ, SUAREZ")

    filtered = matches.copy()
    
    if selected_zone != "Todas":
        filtered = filtered[filtered["Zona"] == selected_zone]

    if selected_complex != "Todos":
        filtered = filtered[filtered["Complejo"] == selected_complex]

    if surname_query.strip():
        query = surname_query.strip().lower()
        filtered = filtered[
            filtered["Pareja 1"].str.lower().str.contains(query, na=False)
            | filtered["Pareja 2"].str.lower().str.contains(query, na=False)
        ]

    st.subheader(selected_label)

    if filtered.empty:
        st.info("No hay partidos para ese filtro en esta categoría.")
        st.stop()

    zones = sorted(filtered["Zona"].dropna().unique().tolist())
    for zone_name in zones:
        st.markdown(f"<div class='zone-title'>{zone_name}</div>", unsafe_allow_html=True)

        zone_matches = filtered[filtered["Zona"] == zone_name].copy()
        if selected_label == "Todas las categorías" and "Categoría" in zone_matches.columns:
            for category_name in sorted(zone_matches["Categoría"].dropna().unique().tolist()):
                st.markdown(f"**{category_name}**")
                render_match_cards(zone_matches[zone_matches["Categoría"] == category_name])
        else:
            render_match_cards(zone_matches)

    with st.expander("Ver tabla"):
        st.dataframe(
            filtered[
                [
                    "Zona",
                    "Categoría",
                    "Día",
                    "Hora",
                    "Complejo",
                    "Código 1",
                    "Pareja 1",
                    "Código 2",
                    "Pareja 2",
                    "Set1 P1",
                    "Set1 P2",
                    "Set2 P1",
                    "Set2 P2",
                    "Set3 P1",
                    "Set3 P2",
                    "Estado",
                ]
            ],
            use_container_width=True,
            hide_index=True,
        )


if __name__ == "__main__":
    main()
